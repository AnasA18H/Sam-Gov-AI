"""
Document Analyzer Service
Extracts text and data from PDF, Word, and Excel documents using hybrid approach:
- Table parsing for structured forms (SF1449, SF30)
- LLM extraction for unstructured text (SOW, amendments)
"""
import re
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import dateutil.parser
from dateutil.parser import ParserError

# Document processing libraries
import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook

# Table extraction (optional - for structured forms)
try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False
    logging.warning("camelot-py not available. Table extraction will use pdfplumber fallback.")

try:
    import tabula
    TABULA_AVAILABLE = True
except ImportError:
    TABULA_AVAILABLE = False
    logging.debug("tabula-py not available. Using camelot or pdfplumber for tables.")

# LLM extraction (optional - for unstructured text)
try:
    from langchain_groq import ChatGroq
    from pydantic import BaseModel, Field
    try:
        from pydantic.v1 import BaseModel as V1BaseModel, Field as V1Field
        PYDANTIC_V1_AVAILABLE = True
    except ImportError:
        # Fallback: use v1 directly if available
        try:
            import pydantic.v1 as pydantic_v1
            V1BaseModel = pydantic_v1.BaseModel
            V1Field = pydantic_v1.Field
            PYDANTIC_V1_AVAILABLE = True
        except ImportError:
            PYDANTIC_V1_AVAILABLE = False
    LANGCHAIN_AVAILABLE = True
    # LangChain 1.x uses with_structured_output instead of create_extraction_chain_pydantic
    EXTRACTION_CHAIN_AVAILABLE = False  # Will use with_structured_output method
except ImportError:
    LANGCHAIN_AVAILABLE = False
    PYDANTIC_V1_AVAILABLE = False
    EXTRACTION_CHAIN_AVAILABLE = False
    logging.warning("LangChain Groq not available. Using regex-based extraction only.")

# NLP (optional - spaCy for advanced classification)
try:
    import spacy
    SPACY_AVAILABLE = True
except ImportError:
    SPACY_AVAILABLE = False
    logging.warning("spaCy not available. Using keyword-based classification only.")

from ..core.config import settings
from ..models.opportunity import SolicitationType
from ..models.clin import CLIN
from ..models.deadline import Deadline

logger = logging.getLogger(__name__)


# Pydantic schema for LLM extraction
# Note: LangChain's create_extraction_chain_pydantic uses Pydantic v1 internally
# So we must use v1 BaseModel for compatibility
if LANGCHAIN_AVAILABLE and PYDANTIC_V1_AVAILABLE:
    class CLINItem(V1BaseModel):
        """
        Pydantic v1 schema for CLIN extraction using LLM.
        
        Extract Contract Line Item Numbers (CLINs) from government solicitation documents.
        Look for structured tables with item numbers, descriptions, quantities, and pricing information.
        """
        item_number: str = V1Field(description="The CLIN or line item number. This is the primary identifier, typically numeric like '0001', '0002', or alphanumeric like '0001AA'. Extract exactly as written in the document.")
        base_item_number: Optional[str] = V1Field(None, description="Base item number or supplementary reference code, often found in adjacent columns. Examples: 'S01', 'AA', 'BASE001'. Only extract if clearly labeled as base item or supplementary code.")
        description: str = V1Field(description="The full description of the product or service. This field is labeled 'Supplies/Services' in many forms. Extract the complete text description, including all details about what is being procured.")
        quantity: Optional[int] = V1Field(None, description="The quantity required as an integer. Extract numeric values only (e.g., 250, 2, 100). Do not include units or text.")
        unit: Optional[str] = V1Field(None, description="The unit of measure for the quantity. Common values: 'Each', 'Lot', 'Set', 'Unit', 'EA'. Extract exactly as written in the document.")
        contract_type: Optional[str] = V1Field(None, description="Contract type or pricing arrangement. Examples: 'Firm Fixed Price', 'Cost Plus Fixed Fee', 'FFP', 'CPFF'. Extract if clearly stated in the CLIN row or table.")
        extended_price: Optional[float] = V1Field(None, description="Extended price or total price for the line item. This may be calculated (quantity × unit price) or explicitly stated. Extract as a numeric value without currency symbols (e.g., 150.00, 18750.00).")
        part_number: Optional[str] = V1Field(None, description="Manufacturer part number if specified. Extract if present in the item details.")
        model_number: Optional[str] = V1Field(None, description="Product model number if specified. Extract if present in the item details.")
        manufacturer: Optional[str] = V1Field(None, description="Manufacturer name if specified. Extract company or brand name if mentioned for this specific CLIN.")
        product_name: Optional[str] = V1Field(None, description="Short product name or title. This is often a condensed version of the description (e.g., 'Stack and Rack Carts', 'Desktop Computers'). Extract if clearly distinguishable from the full description.")
elif LANGCHAIN_AVAILABLE:
    # Fallback: define empty class if v1 not available
    class CLINItem:
        pass


class DocumentAnalyzer:
    """Analyzer for extracting text and structured data from solicitation documents"""
    
    @staticmethod
    def _clean_text(text: str) -> str:
        """
        Clean extracted text by removing PDF encoding artifacts and other garbage
        Removes (cid:XXX) patterns which are PDF character encoding artifacts
        Also removes corrupted patterns like semicolon-separated garbage text
        """
        if not text:
            return ""
        
        # Remove PDF encoding artifacts (cid:XXX) where XXX is any number
        text = re.sub(r'\(cid:\d+\)', '', text)
        
        # Remove corrupted patterns: semicolon-separated alphanumeric garbage
        # Pattern: ; followed by alphanumeric chars, <, >, =, :, etc. ending with ;
        # Example: ;9<746Q57BG69B<;4C8=B51A75BA74:5;38<=5914:A==93A;
        text = re.sub(r';[A-Z0-9<>=:]+(?:;[A-Z0-9<>=:]+)+;', '', text)
        
        # Remove lines that are mostly garbage (high ratio of special chars to letters)
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            # Skip lines that are mostly special characters and numbers with few actual words
            if len(line.strip()) > 0:
                # Count actual letters (a-z, A-Z)
                letter_count = len(re.findall(r'[a-zA-Z]', line))
                # Count total alphanumeric chars
                alnum_count = len(re.findall(r'[a-zA-Z0-9]', line))
                total_chars = len(line.strip())
                
                # If line has very few letters relative to total chars, or is mostly special chars
                if total_chars > 10 and (letter_count < 3 or (alnum_count / total_chars) < 0.3):
                    # This is likely garbage, skip it
                    continue
            cleaned_lines.append(line)
        text = '\n'.join(cleaned_lines)
        
        # Remove other common PDF artifacts
        text = re.sub(r'[^\x20-\x7E\n\r\t]', '', text)  # Remove non-printable chars except newlines/tabs
        text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
        
        return text.strip()
    
    # Product-related keywords
    PRODUCT_KEYWORDS = [
        'product', 'item', 'manufacturer', 'part number', 'model number', 'catalog',
        'equipment', 'supply', 'material', 'component', 'hardware', 'device',
        'unit', 'piece', 'quantity', 'delivery', 'ship', 'furnish', 'provide'
    ]
    
    # Service-related keywords
    SERVICE_KEYWORDS = [
        'service', 'work', 'task', 'perform', 'maintain', 'repair', 'install',
        'support', 'consult', 'training', 'hours', 'labor', 'personnel',
        'contractor', 'vendor', 'timeline', 'schedule', 'duration', 'period'
    ]
    
    # Document type classification patterns
    DOCUMENT_TYPE_PATTERNS = {
        'sf1449': [r'SF\s*1449', r'Standard Form\s*1449', r'Form\s*1449'],
        'sf30': [r'SF\s*30', r'Standard Form\s*30', r'Amendment'],
        'sow': [r'Statement\s+of\s+Work', r'SOW', r'Scope\s+of\s+Work'],
        'amendment': [r'Amendment', r'Modification', r'Change\s+Order'],
    }
    
    # CLIN patterns (for regex fallback)
    # CLINs are typically: 0001, 0002, 1, 2, or with suffix like 0001AA, 0001A
    # NOT short numbers with letters like 45B (those are usually base items)
    CLIN_PATTERNS = [
        r'CLIN\s*(\d{3,}[A-Z]*)',  # At least 3 digits (0001, 0002, etc.)
        r'CLIN\s*(\d{1,2}[A-Z]{2,})',  # 1-2 digits with 2+ letters (0001AA)
        r'Contract\s+Line\s+Item\s+Number\s*(\d{3,}[A-Z]*)',
        r'Line\s+Item\s+(\d{3,}[A-Z]*)',
        r'Item\s+No\.?\s*(\d{3,}[A-Z]*)',  # Item No. with at least 3 digits
    ]
    
    # Deadline patterns
    DATE_PATTERNS = [
        r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})',
        r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),?\s+(\d{4})',
        r'(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})',
    ]
    
    TIME_PATTERNS = [
        r'(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)',
        r'(\d{1,2})\s*(AM|PM|am|pm)',
        r'(\d{4})\s*hours?',
    ]
    
    DEADLINE_KEYWORDS = [
        'due', 'deadline', 'submission', 'offer', 'proposal', 'response',
        'quote', 'bid', 'close', 'receipt', 'must be received'
    ]
    
    def __init__(self):
        """Initialize the document analyzer"""
        self.nlp = None
        self.llm = None
        
        # Initialize spaCy (optional)
        if SPACY_AVAILABLE:
            try:
                self.nlp = spacy.load("en_core_web_sm")
                logger.info("spaCy model loaded successfully")
            except OSError:
                logger.warning("spaCy model 'en_core_web_sm' not found. Using keyword-based classification only.")
        
        # Initialize LLM (optional - requires GROQ_API_KEY)
        if LANGCHAIN_AVAILABLE and settings.GROQ_API_KEY:
            try:
                self.llm = ChatGroq(
                    model=settings.GROQ_MODEL,
                    temperature=0,
                    groq_api_key=settings.GROQ_API_KEY
                )
                logger.info(f"Groq LLM initialized: {settings.GROQ_MODEL}")
            except Exception as e:
                logger.warning(f"Failed to initialize Groq LLM: {str(e)}. Using regex-based extraction only.")
                self.llm = None
        elif LANGCHAIN_AVAILABLE and not settings.GROQ_API_KEY:
            logger.warning("GROQ_API_KEY not set. LLM extraction disabled. Using regex-based extraction only.")
    
    def classify_document_type(self, file_path: Path, text: str) -> str:
        """
        Classify document type to route to correct extraction method
        
        Returns:
            Document type: 'sf1449', 'sf30', 'sow', 'amendment', or 'unknown'
        """
        # Check filename first
        filename_lower = file_path.name.lower()
        for doc_type, patterns in self.DOCUMENT_TYPE_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, filename_lower, re.IGNORECASE):
                    logger.info(f"Document classified as {doc_type} based on filename")
                    return doc_type
        
        # Check text content
        text_lower = text[:2000].lower()  # Check first 2000 chars
        for doc_type, patterns in self.DOCUMENT_TYPE_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, text_lower, re.IGNORECASE):
                    logger.info(f"Document classified as {doc_type} based on content")
                    return doc_type
        
        return 'unknown'
    
    def extract_text(self, file_path: str) -> str:
        """
        Extract text from a document file
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Extracted text content
        """
        file_path_obj = Path(file_path)
        
        # Handle relative paths (for Digital Ocean compatibility)
        if not file_path_obj.is_absolute():
            project_root = Path(settings.PROJECT_ROOT)
            abs_path = project_root / file_path
            if abs_path.exists():
                file_path_obj = abs_path
            elif hasattr(settings, 'STORAGE_BASE_PATH'):
                storage_base = Path(settings.STORAGE_BASE_PATH)
                abs_path = storage_base.parent / file_path if 'backend/data' in file_path else storage_base / file_path
                if abs_path.exists():
                    file_path_obj = abs_path
        
        if not file_path_obj.exists():
            raise FileNotFoundError(f"Document file not found: {file_path}")
        
        file_ext = file_path_obj.suffix.lower()
        
        try:
            if file_ext == '.txt':
                # TXT files are already text - just read them directly
                logger.info(f"Reading TXT file directly: {file_path_obj.name}")
                with open(file_path_obj, 'r', encoding='utf-8') as f:
                    text = f.read()
                # Skip the header if it's from our extraction (starts with "Source URL:")
                if text.startswith("Source URL:"):
                    # Find where the actual content starts (after the separator line)
                    lines = text.split('\n')
                    content_start = 0
                    for i, line in enumerate(lines):
                        if line.startswith("=" * 80) or line.startswith("=" * 40):
                            content_start = i + 1
                            break
                    if content_start > 0:
                        text = '\n'.join(lines[content_start:])
                logger.info(f"Read {len(text)} characters from TXT file")
                return text
            elif file_ext == '.pdf':
                return self._extract_from_pdf(file_path_obj)
            elif file_ext in ['.doc', '.docx']:
                return self._extract_from_word(file_path_obj)
            elif file_ext in ['.xls', '.xlsx']:
                return self._extract_from_excel(file_path_obj)
            else:
                logger.warning(f"Unsupported file type: {file_ext}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}", exc_info=True)
            return ""
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF file"""
        text_parts = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        # Clean PDF encoding artifacts
                        cleaned_text = self._clean_text(page_text)
                        if cleaned_text:
                            text_parts.append(cleaned_text)
            return "\n\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from PDF {file_path}: {str(e)}")
            return ""
    
    def _extract_from_word(self, file_path: Path) -> str:
        """Extract text from Word document"""
        try:
            doc = DocxDocument(file_path)
            text_parts = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_parts.append(paragraph.text)
            return "\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from Word doc {file_path}: {str(e)}")
            return ""
    
    def _extract_from_excel(self, file_path: Path) -> str:
        """Extract text from Excel file"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        sheet_text.append(row_text)
                if sheet_text:
                    text_parts.append(f"\n--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_text))
            return "\n\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from Excel {file_path}: {str(e)}")
            return ""
    
    @staticmethod
    def _is_valid_clin_number(cell_value: str) -> bool:
        """
        Check if cell value looks like a valid CLIN number.
        CLINs are typically: 0001, 0002, 1, 2, or with suffix like 0001AA, 0001A
        NOT short numbers with single letters like 45B (those are usually base items)
        """
        if not cell_value or len(cell_value.strip()) == 0:
            return False
        # Match: at least 3 digits, OR 1-2 digits with 2+ letters (like 1AA, 2AB)
        return bool(re.match(r'^(\d{3,}[A-Z]*|\d{1,2}[A-Z]{2,})$', cell_value.strip()))
    
    def _extract_clins_from_table(self, file_path: Path) -> List[Dict]:
        """
        Extract CLINs from PDF tables using camelot-py or tabula-py
        Best for structured forms like SF1449, SF30
        
        Returns:
            List of CLIN dictionaries
        """
        if file_path.suffix.lower() != '.pdf':
            return []
        
        clins = []
        
        # Try camelot first
        if CAMELOT_AVAILABLE:
            try:
                tables = camelot.read_pdf(str(file_path), pages='all', flavor='lattice')
                logger.info(f"Found {len(tables)} tables with camelot")
                
                for table in tables:
                    df = table.df
                    # Look for CLIN column (usually first column)
                    for idx, row in df.iterrows():
                        # Check if first column looks like a CLIN number
                        first_cell = str(row.iloc[0]).strip() if len(row) > 0 else ""
                        if self._is_valid_clin_number(first_cell):
                            raw_description = str(row.iloc[1]) if len(row) > 1 else None
                            cleaned_description = self._clean_text(raw_description) if raw_description else None
                            
                            # Skip CLINs with garbage descriptions
                            if cleaned_description:
                                desc_clean = cleaned_description.strip()
                                # Skip if too short
                                if len(desc_clean) < 3:
                                    continue
                                # Skip if no letters (only numbers/symbols)
                                if not any(c.isalpha() for c in desc_clean):
                                    continue
                                # Skip if it looks like corrupted encoding (semicolon-separated garbage)
                                if re.search(r'^;[A-Z0-9<>=:]+(?:;[A-Z0-9<>=:]+)+;?$', desc_clean):
                                    continue
                                # Skip if mostly special characters (less than 30% letters/numbers)
                                letter_count = len(re.findall(r'[a-zA-Z]', desc_clean))
                                alnum_count = len(re.findall(r'[a-zA-Z0-9]', desc_clean))
                                if len(desc_clean) > 10 and (letter_count < 3 or (alnum_count / len(desc_clean)) < 0.3):
                                continue
                            
                            clin_data = {
                                'clin_number': first_cell,
                                'base_item_number': None,
                                'product_description': cleaned_description,
                                'quantity': None,
                                'unit_of_measure': None,
                                'contract_type': None,
                                'extended_price': None,
                                'part_number': None,
                                'model_number': None,
                                'manufacturer_name': None,
                            }
                            
                            # Try to extract data from other columns
                            for col_idx, col in enumerate(df.columns):
                                cell_value = str(row[col]).strip() if col_idx < len(row) else ""
                                cell_lower = cell_value.lower()
                                
                                # Extract quantity
                                if not clin_data['quantity']:
                                    qty_match = re.search(r'(\d+(?:\.\d+)?)', cell_lower)
                                    if qty_match:
                                        try:
                                            clin_data['quantity'] = float(qty_match.group(1))
                                        except ValueError:
                                            pass
                                
                                # Extract unit of measure
                                if not clin_data['unit_of_measure']:
                                    unit_patterns = ['each', 'lot', 'set', 'unit', 'piece', 'ea', 'lot']
                                    if any(unit in cell_lower for unit in unit_patterns):
                                        clin_data['unit_of_measure'] = cell_value
                                
                                # Extract base item number (usually alphanumeric like S01, AA, etc.)
                                if not clin_data['base_item_number']:
                                    base_item_match = re.search(r'\b([A-Z]{1,3}\d{1,3}|[A-Z]{2,4})\b', cell_value)
                                    if base_item_match and base_item_match.group(1) != first_cell:
                                        clin_data['base_item_number'] = base_item_match.group(1)
                                
                                # Extract contract type
                                if not clin_data['contract_type']:
                                    contract_patterns = ['firm fixed price', 'cost plus', 'time and materials', 'fixed price', 'ffp', 'cpff']
                                    if any(pattern in cell_lower for pattern in contract_patterns):
                                        clin_data['contract_type'] = cell_value
                                
                                # Extract extended price (usually contains $ or currency symbols)
                                if not clin_data['extended_price']:
                                    price_match = re.search(r'[\$]?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', cell_value)
                                    if price_match:
                                        try:
                                            price_str = price_match.group(1).replace(',', '')
                                            clin_data['extended_price'] = float(price_str)
                                        except ValueError:
                                            pass
                            
                            clins.append(clin_data)
            except Exception as e:
                logger.warning(f"camelot table extraction failed: {str(e)}")
        
        # Fallback to pdfplumber tables
        if not clins:
            try:
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        tables = page.extract_tables()
                        for table in tables:
                            if len(table) > 1:  # Has header row
                                for row in table[1:]:  # Skip header
                                    if row and len(row) > 0:
                                        first_cell = str(row[0]).strip() if row[0] else ""
                                        if self._is_valid_clin_number(first_cell):
                                            raw_description = str(row[1]) if len(row) > 1 else None
                                            cleaned_description = self._clean_text(raw_description) if raw_description else None
                                            
                                            # Validate description - if corrupted, set to None (we'll try to find it later)
                                            if cleaned_description:
                                                desc_clean = cleaned_description.strip()
                                                # Check if description is valid
                                                is_valid = True
                                                # Skip if too short
                                                if len(desc_clean) < 3:
                                                    is_valid = False
                                                # Skip if no letters (only numbers/symbols)
                                                elif not any(c.isalpha() for c in desc_clean):
                                                    is_valid = False
                                                # Skip if it looks like corrupted encoding (semicolon-separated garbage)
                                                elif re.search(r'^;[A-Z0-9<>=:]+(?:;[A-Z0-9<>=:]+)+;?$', desc_clean):
                                                    is_valid = False
                                                # Skip if mostly special characters (less than 30% letters/numbers)
                                                else:
                                                    letter_count = len(re.findall(r'[a-zA-Z]', desc_clean))
                                                    alnum_count = len(re.findall(r'[a-zA-Z0-9]', desc_clean))
                                                    if len(desc_clean) > 10 and (letter_count < 3 or (alnum_count / len(desc_clean)) < 0.3):
                                                        is_valid = False
                                                
                                                if not is_valid:
                                                    cleaned_description = None
                                                    logger.debug(f"CLIN {first_cell} has corrupted description, will try to find it from document text")
                                            
                                            clin_data = {
                                                'clin_number': first_cell,
                                                'base_item_number': None,
                                                'product_description': cleaned_description,
                                                'quantity': None,
                                                'unit_of_measure': None,
                                                'contract_type': None,
                                                'extended_price': None,
                                            }
                                            
                                            # Try to extract additional fields from other columns
                                            for col_idx in range(2, len(row)):
                                                cell_value = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] else ""
                                                cell_lower = cell_value.lower()
                                                
                                                # Extract quantity
                                                if not clin_data['quantity']:
                                                    qty_match = re.search(r'(\d+(?:\.\d+)?)', cell_lower)
                                                    if qty_match:
                                                        try:
                                                            clin_data['quantity'] = float(qty_match.group(1))
                                                        except ValueError:
                                                            pass
                                                
                                                # Extract unit
                                                if not clin_data['unit_of_measure']:
                                                    if any(unit in cell_lower for unit in ['each', 'lot', 'set', 'unit', 'ea']):
                                                        clin_data['unit_of_measure'] = cell_value
                                                
                                                # Extract base item number
                                                if not clin_data['base_item_number']:
                                                    base_match = re.search(r'\b([A-Z]{1,3}\d{1,3}|[A-Z]{2,4})\b', cell_value)
                                                    if base_match and base_match.group(1) != first_cell:
                                                        clin_data['base_item_number'] = base_match.group(1)
                                                
                                                # Extract contract type
                                                if not clin_data['contract_type']:
                                                    if any(pattern in cell_lower for pattern in ['firm fixed price', 'cost plus', 'ffp']):
                                                        clin_data['contract_type'] = cell_value
                                                
                                                # Extract extended price
                                                if not clin_data['extended_price']:
                                                    price_match = re.search(r'[\$]?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', cell_value)
                                                    if price_match:
                                                        try:
                                                            price_str = price_match.group(1).replace(',', '')
                                                            clin_data['extended_price'] = float(price_str)
                                                        except ValueError:
                                                            pass
                                            
                                            clins.append(clin_data)
            except Exception as e:
                logger.warning(f"pdfplumber table extraction failed: {str(e)}")
        
        return clins
    
    def _extract_clins_with_llm(self, text: str) -> List[Dict]:
        """
        Extract CLINs using LLM (LangChain + Groq with Llama models) with batch processing
        Best for unstructured text like SOW, amendments
        
        Returns:
            List of CLIN dictionaries
        """
        if not self.llm or not LANGCHAIN_AVAILABLE or not PYDANTIC_V1_AVAILABLE:
            if not PYDANTIC_V1_AVAILABLE:
                logger.warning("Pydantic v1 not available. LLM extraction requires Pydantic v1 for LangChain compatibility.")
            return []
        
        try:
            # Skip Q&A documents
            text_lower = text.lower()
            qa_indicators = ['could the government', 'question', 'government clarify', 'q&a']
            if any(indicator in text_lower[:500] for indicator in qa_indicators):
                logger.debug("Skipping LLM extraction from Q&A document")
                return []
            
            # Clean text before LLM processing
            cleaned_text = self._clean_text(text)
            
            if not cleaned_text.strip():
                logger.debug("Cleaned text is empty, skipping LLM extraction")
                return []
            
            # Process entire document at once (send full document, not split into batches)
            # This ensures we send one document at a time, sequentially
            # Use reasonable context window - send up to 15000 chars per document to avoid payload size limits
            max_document_size = 15000
            document_text = cleaned_text[:max_document_size]
            
            if not document_text.strip():
                return []
            
            logger.info(f"Processing document for LLM extraction ({len(document_text)} chars)")
            
            # Simple, general prompt - process entire document
            document_instruction = """Process this document for CLINs (Contract Line Item Numbers).

If you find CLINs in THIS document, return them. If none found, return NONE.

Extract: item_number, description (readable English only), quantity, unit, extended_price, and other fields if available.
Ignore corrupted text and encoding artifacts. Process the entire document.

DOCUMENT TEXT:
"""
            
            # Process document (one document at a time, no batching)
            import time
            max_retries = 3
            retry_delay = 2  # seconds
            for attempt in range(max_retries):
                try:
                    document_prompt = document_instruction + document_text
            
                    # Use LangChain 1.x API: with_structured_output
                    try:
                        # Try function_calling first (better Groq compatibility), fallback to json_schema
                        try:
                            structured_llm = self.llm.with_structured_output(CLINItem, method="function_calling")
                        except Exception as method_error:
                            logger.debug(f"function_calling failed, trying json_schema: {str(method_error)}")
                            # Fallback to json_schema method
                            structured_llm = self.llm.with_structured_output(CLINItem, method="json_schema")
                        result = structured_llm.invoke(document_prompt)
                    except Exception as e:
                        # Check if it's a rate limit error
                        error_str = str(e).lower()
                        if '429' in error_str or 'rate limit' in error_str or 'too many requests' in error_str:
                            if attempt < max_retries - 1:
                                wait_time = retry_delay * (attempt + 1)
                                logger.warning(f"Rate limit hit, waiting {wait_time}s before retry {attempt + 2}/{max_retries}")
                                time.sleep(wait_time)
                                continue
                            else:
                                logger.error(f"Rate limit exceeded after {max_retries} attempts")
                                return []
                        
                        # Log the actual error for debugging
                        logger.warning(f"with_structured_output failed: {str(e)}")
                        
                        # Fallback: try direct invoke with JSON parsing
                        try:
                            logger.debug(f"Trying direct invoke with JSON parsing")
                            response = self.llm.invoke(document_prompt + "\n\nReturn the CLINs as a JSON array.")
                            # Try to parse response content
                            if hasattr(response, 'content'):
                                import json
                                try:
                                    # Try to parse as JSON
                                    result = json.loads(response.content)
                    if isinstance(result, list):
                                        result = [CLINItem(**item) if isinstance(item, dict) else item for item in result]
                                except:
                                    result = response.content
                            else:
                                result = str(response)
                        except Exception as fallback_error:
                            logger.error(f"All extraction methods failed: {str(fallback_error)}")
                            result = None
                    
                    # Extract results from response
                    clins = []
                    if isinstance(result, list):
                        clins = result
                    elif isinstance(result, CLINItem):
                        clins = [result]
                    elif isinstance(result, dict):
                        # Check if it's a single CLINItem dict
                        if 'item_number' in result:
                            clins = [result]
                        else:
                            # Try to find list in dict values
                            for value in result.values():
                                if isinstance(value, list):
                                    clins = value
                                    break
                                elif isinstance(value, CLINItem):
                                    clins = [value]
                                    break
                    
                    if clins:
                        logger.info(f"Found {len(clins)} CLINs in document")
                    else:
                        logger.debug("No CLINs found in document")
                    
                    # Success - break out of retry loop
                    break
                        
                except Exception as doc_error:
                    error_str = str(doc_error).lower()
                    # Check if it's a rate limit or temporary error
                    if ('429' in error_str or 'rate limit' in error_str or 'too many requests' in error_str) and attempt < max_retries - 1:
                        wait_time = retry_delay * (attempt + 1)
                        logger.warning(f"Rate limit error, waiting {wait_time}s before retry {attempt + 2}/{max_retries}")
                        time.sleep(wait_time)
                        continue
                    elif attempt == max_retries - 1:
                        logger.error(f"Failed to process document after {max_retries} attempts: {str(doc_error)}")
                        return []
                    else:
                        logger.warning(f"Error processing document (attempt {attempt + 1}/{max_retries}): {str(doc_error)}")
                        time.sleep(retry_delay)
                        continue
            
            # Convert results to our format
            if clins:
                converted_clins = self._convert_llm_results_to_dicts(clins)
                if len(converted_clins) != len(clins):
                    logger.warning(f"CLIN conversion filtered out {len(clins) - len(converted_clins)} CLIN(s) (found {len(clins)}, kept {len(converted_clins)})")
                return converted_clins
            else:
                return []
            
            except Exception as e:
            logger.error(f"LLM extraction failed: {str(e)}")
                logger.debug(f"Error details: {type(e).__name__}", exc_info=True)
            return []
            
    def _convert_llm_results_to_dicts(self, results: List) -> List[Dict]:
        """Convert LLM results (CLINItem objects or dicts) to our standard dict format"""
            clins = []
        if not isinstance(results, list):
            return clins
        
                for item in results:
                    try:
                        # Handle both CLINItem objects and dicts
                        if isinstance(item, CLINItem):
                            # Extract from Pydantic model
                            item_number = item.item_number
                            description = self._clean_text(item.description) if item.description else None
                            product_name = self._clean_text(item.product_name) if item.product_name else None
                            quantity = float(item.quantity) if item.quantity else None
                            unit = item.unit
                            base_item_number = getattr(item, 'base_item_number', None)
                            contract_type = getattr(item, 'contract_type', None)
                            extended_price = float(item.extended_price) if getattr(item, 'extended_price', None) else None
                            part_number = item.part_number
                            model_number = item.model_number
                            manufacturer = item.manufacturer
                        elif isinstance(item, dict):
                            # Extract from dict (in case LLM returns dict instead of model)
                            item_number = item.get('item_number', '')
                            description = self._clean_text(item.get('description', '')) if item.get('description') else None
                            product_name = self._clean_text(item.get('product_name', '')) if item.get('product_name') else None
                            quantity = float(item['quantity']) if item.get('quantity') else None
                            unit = item.get('unit')
                            base_item_number = item.get('base_item_number')
                            contract_type = item.get('contract_type')
                            extended_price = float(item['extended_price']) if item.get('extended_price') else None
                            part_number = item.get('part_number')
                            model_number = item.get('model_number')
                            manufacturer = item.get('manufacturer')
                        else:
                            continue
                        
                # Validate CLIN number (must be valid format)
                if not item_number:
                    logger.debug(f"Skipping CLIN with empty item_number")
                            continue
                if not self._is_valid_clin_number(str(item_number)):
                    logger.debug(f"Skipping CLIN with invalid item_number format: '{item_number}'")
                            continue
                        
                # Validate description (must be readable, not corrupted)
                if description:
                    desc_clean = description.strip()
                    if len(desc_clean) < 3:
                        description = None
                    elif not any(c.isalpha() for c in desc_clean):
                        description = None
                    elif re.search(r'^;[A-Z0-9<>=:]+(?:;[A-Z0-9<>=:]+)+;?$', desc_clean):
                        description = None
                    else:
                        letter_count = len(re.findall(r'[a-zA-Z]', desc_clean))
                        alnum_count = len(re.findall(r'[a-zA-Z0-9]', desc_clean))
                        if len(desc_clean) > 10 and (letter_count < 3 or (alnum_count / len(desc_clean)) < 0.3):
                            description = None
                
                # Only add if we have at least a CLIN number and some valid data
                if item_number:
                        clin_data = {
                            'clin_number': str(item_number).strip(),
                            'base_item_number': str(base_item_number).strip() if base_item_number else None,
                            'product_description': description,
                            'quantity': quantity,
                            'unit_of_measure': str(unit).strip() if unit else None,
                            'contract_type': str(contract_type).strip() if contract_type else None,
                            'extended_price': extended_price,
                        'product_name': product_name,
                        'manufacturer_name': str(manufacturer).strip() if manufacturer else None,
                            'part_number': str(part_number).strip() if part_number else None,
                            'model_number': str(model_number).strip() if model_number else None,
                        }
                        clins.append(clin_data)
                    
                    except Exception as item_error:
                logger.warning(f"Error converting LLM result item: {str(item_error)}")
                        continue
            
            return clins
    
    def extract_clins(self, text: str, file_path: Optional[Path] = None) -> List[Dict]:
        """
        Extract CLINs using ONLY AI/LLM extraction.
        No table extraction, no regex fallback - AI only.
        
        Args:
            text: Document text content
            file_path: Optional path to document file (not used, kept for compatibility)
            
        Returns:
            List of CLIN dictionaries with extracted data
        """
        # Skip Q&A documents early
        text_lower = text.lower()
        qa_patterns = [
            r'could\s+the\s+government',
            r'question\s*\d+',
            r'government\s+clarif',
            r'q&a',
            r'q/a',
        ]
        if any(re.search(pattern, text_lower[:500]) for pattern in qa_patterns):
            logger.debug("Skipping CLIN extraction from Q&A document")
            return []
        
        # ONLY use LLM extraction - no fallbacks
        if not self.llm:
            logger.warning("LLM not available. Cannot extract CLINs (AI-only mode).")
            return []
        
        # Classify document type for logging
        doc_type = 'unknown'
        if file_path:
            doc_type = self.classify_document_type(file_path, text)
        
        logger.info(f"Using AI/LLM extraction for {doc_type} (AI-only mode)")
            llm_clins = self._extract_clins_with_llm(text)
        
            if llm_clins:
            logger.info(f"AI extraction found {len(llm_clins)} CLINs")
        else:
            logger.info("AI extraction found 0 CLINs")
        
                return llm_clins
        
    def _find_clin_description_in_text(self, text: str, clin_number: str) -> Optional[str]:
        """
        Search for CLIN number in text and extract nearby description
        Used as fallback when table extraction finds CLIN but description is corrupted
        """
        # Search for CLIN number patterns in text
        patterns = [
            rf'CLIN\s+{re.escape(clin_number)}[:\s]+([^.\n]{20,500})',
            rf'Item\s+{re.escape(clin_number)}[:\s]+([^.\n]{20,500})',
            rf'Line\s+Item\s+{re.escape(clin_number)}[:\s]+([^.\n]{20,500})',
            rf'{re.escape(clin_number)}[:\s]+([^.\n]{20,500})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                desc = match.group(1).strip()
                # Clean and validate the description
                desc = self._clean_text(desc)
                # Check if it's valid (not corrupted)
                if desc and len(desc) >= 10:
                    letter_count = len(re.findall(r'[a-zA-Z]', desc))
                    if letter_count >= 5:  # At least 5 letters
                        # Take first sentence or first 200 chars
                        sentences = re.split(r'[.!?]\s+', desc)
                        if sentences:
                            result = sentences[0][:200].strip()
                            if len(result) >= 10:
                                return result
        
        # If no direct match, look for product/service descriptions near the CLIN
        # Search for common description patterns
        clin_pos = text.find(clin_number)
        if clin_pos != -1:
            # Look in a window around the CLIN number
            start = max(0, clin_pos - 500)
            end = min(len(text), clin_pos + 1000)
            window = text[start:end]
            
            # Look for description patterns
            desc_patterns = [
                r'Supplies/Services[:\s]+([^.\n]{20,300})',
                r'Description[:\s]+([^.\n]{20,300})',
                r'Item\s+Description[:\s]+([^.\n]{20,300})',
            ]
            
            for pattern in desc_patterns:
                match = re.search(pattern, window, re.IGNORECASE)
                if match:
                    desc = match.group(1).strip()
                    desc = self._clean_text(desc)
                    if desc and len(desc) >= 10:
                        letter_count = len(re.findall(r'[a-zA-Z]', desc))
                        if letter_count >= 5:
                            return desc[:200].strip()
        
        return None
    
    def _extract_clins_regex(self, text: str) -> List[Dict]:
        """Fallback regex-based CLIN extraction (original method)"""
        clins = []
        
        # Find all CLIN numbers in the text
        clin_matches = []
        for pattern in self.CLIN_PATTERNS:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                clin_num = match.group(1)
                position = match.start()
                clin_matches.append((clin_num, position, match))
        
        # Sort by position in document
        clin_matches.sort(key=lambda x: x[1])
        
        # Extract details for each CLIN
        for i, (clin_num, position, match) in enumerate(clin_matches):
            start_pos = position
            end_pos = clin_matches[i + 1][1] if i + 1 < len(clin_matches) else min(position + 2000, len(text))
            clin_text = text[start_pos:end_pos]
            
            # Skip Q&A text
            if any(re.search(pattern, clin_text[:200].lower()) for pattern in [
                r'could\s+the\s+government', r'question\s*\d+', r'government\s+clarif'
            ]):
                logger.debug(f"Skipping CLIN {clin_num} - appears to be Q&A text")
                continue
            
            # Extract CLIN details using regex
            clin_data = {
                'clin_number': clin_num,
                'base_item_number': None,
                'product_name': None,
                'product_description': None,
                'manufacturer_name': None,
                'part_number': None,
                'model_number': None,
                'contract_type': None,
                'extended_price': None,
                'quantity': None,
                'unit_of_measure': None,
                'service_description': None,
            }
            
            # Extract quantity
            qty_patterns = [r'quantity[:\s]+(\d+(?:\.\d+)?)', r'qty[:\s]+(\d+(?:\.\d+)?)']
            for pattern in qty_patterns:
                match = re.search(pattern, clin_text, re.IGNORECASE)
                if match:
                    try:
                        clin_data['quantity'] = float(match.group(1))
                    except ValueError:
                        pass
                    break
            
            # Extract unit of measure
            unit_patterns = [r'unit[:\s]+([A-Za-z]+)', r'uom[:\s]+([A-Za-z]+)', r'\b(each|lot|set|unit|piece|ea)\b']
            for pattern in unit_patterns:
                match = re.search(pattern, clin_text, re.IGNORECASE)
                if match:
                    clin_data['unit_of_measure'] = match.group(1).capitalize()
                    break
            
            # Extract base item number
            base_item_match = re.search(r'base\s+item[:\s]+([A-Z]{1,3}\d{1,3}|[A-Z]{2,4})', clin_text, re.IGNORECASE)
            if base_item_match:
                clin_data['base_item_number'] = base_item_match.group(1)
            
            # Extract contract type
            contract_patterns = [
                r'contract\s+type[:\s]+([^,\n]+)',
                r'(firm\s+fixed\s+price|cost\s+plus|time\s+and\s+materials|fixed\s+price|ffp|cpff)',
            ]
            for pattern in contract_patterns:
                match = re.search(pattern, clin_text, re.IGNORECASE)
                if match:
                    clin_data['contract_type'] = match.group(1).strip()
                    break
            
            # Extract extended price
            price_patterns = [
                r'extended\s+price[:\s]*[\$]?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                r'total[:\s]*[\$]?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
            ]
            for pattern in price_patterns:
                match = re.search(pattern, clin_text, re.IGNORECASE)
                if match:
                    try:
                        price_str = match.group(1).replace(',', '')
                        clin_data['extended_price'] = float(price_str)
                    except ValueError:
                        pass
                    break
            
            # Extract description (first sentence after CLIN)
            sentences = re.split(r'[.!?]\s+', clin_text[:1000])
            if sentences and len(sentences) > 1:
                clin_data['product_description'] = sentences[1][:500]
            
            clins.append(clin_data)
        
        logger.info(f"Regex extraction found {len(clins)} CLINs")
        return clins
    
    def classify_solicitation_type(self, text: str, title: Optional[str] = None, description: Optional[str] = None) -> Tuple[SolicitationType, float]:
        """
        Classify solicitation type as product, service, or both
        
        Args:
            text: Extracted text from documents
            title: Opportunity title (optional)
            description: Opportunity description (optional)
            
        Returns:
            Tuple of (SolicitationType, confidence_score)
        """
        combined_text = " ".join([title or "", description or "", text]).lower()
        
        product_matches = sum(1 for keyword in self.PRODUCT_KEYWORDS if keyword in combined_text)
        service_matches = sum(1 for keyword in self.SERVICE_KEYWORDS if keyword in combined_text)
        
        total_matches = product_matches + service_matches
        if total_matches == 0:
            return SolicitationType.UNKNOWN, 0.0
        
        product_ratio = product_matches / total_matches
        service_ratio = service_matches / total_matches
        
        if product_ratio > 0.6:
            classification = SolicitationType.PRODUCT
            confidence = product_ratio
        elif service_ratio > 0.6:
            classification = SolicitationType.SERVICE
            confidence = service_ratio
        else:
            classification = SolicitationType.BOTH
            confidence = 1.0 - abs(product_ratio - service_ratio)
        
        if self.nlp and len(combined_text) > 100:
            try:
                doc = self.nlp(combined_text[:5000])
                nouns = [token.text.lower() for token in doc if token.pos_ == "NOUN"]
                verbs = [token.text.lower() for token in doc if token.pos_ == "VERB"]
                
                product_nouns = sum(1 for noun in nouns if any(kw in noun for kw in self.PRODUCT_KEYWORDS))
                service_verbs = sum(1 for verb in verbs if any(kw in verb for kw in self.SERVICE_KEYWORDS))
                
                if product_nouns > 0 or service_verbs > 0:
                    if product_nouns > service_verbs * 2:
                        classification = SolicitationType.PRODUCT
                        confidence = min(0.95, product_nouns / (product_nouns + service_verbs))
                    elif service_verbs > product_nouns * 2:
                        classification = SolicitationType.SERVICE
                        confidence = min(0.95, service_verbs / (product_nouns + service_verbs))
                    else:
                        classification = SolicitationType.BOTH
                        confidence = 0.85
            except Exception as e:
                logger.warning(f"spaCy classification failed: {str(e)}")
        
        logger.info(f"Classification: {classification.value}, confidence: {confidence:.2f}")
        return classification, confidence
    
    def extract_deadlines(self, text: str) -> List[Dict]:
        """
        Extract deadlines from document text
        
        Args:
            text: Document text content
            
        Returns:
            List of deadline dictionaries
        """
        deadlines = []
        
        deadline_pattern = r'(?:' + '|'.join(self.DEADLINE_KEYWORDS) + r')[:\s]*([^.\n]{10,200})'
        deadline_matches = re.finditer(deadline_pattern, text, re.IGNORECASE)
        
        for match in deadline_matches:
            deadline_context = match.group(0)
            
            for date_pattern in self.DATE_PATTERNS:
                date_match = re.search(date_pattern, deadline_context)
                if date_match:
                    try:
                        date_str = date_match.group(0)
                        parsed_date = dateutil.parser.parse(date_str, fuzzy=True, default=datetime(1900, 1, 1))
                        
                        time_str = None
                        timezone_str = None
                        
                        for time_pattern in self.TIME_PATTERNS:
                            time_match = re.search(time_pattern, deadline_context)
                            if time_match:
                                time_str = time_match.group(0)
                                break
                        
                        tz_patterns = [r'\b(EST|EDT|CST|CDT|MST|MDT|PST|PDT|UTC)\b']
                        for tz_pattern in tz_patterns:
                            tz_match = re.search(tz_pattern, deadline_context, re.IGNORECASE)
                            if tz_match:
                                timezone_str = tz_match.group(1).upper()
                                break
                        
                        deadline_type = "submission"
                        if any(kw in deadline_context.lower() for kw in ['question', 'inquiry']):
                            deadline_type = "questions_due"
                        elif any(kw in deadline_context.lower() for kw in ['quote', 'bid', 'proposal']):
                            deadline_type = "submission"
                        elif any(kw in deadline_context.lower() for kw in ['offer']):
                            deadline_type = "offers_due"
                        
                        deadlines.append({
                            'due_date': parsed_date,
                            'due_time': time_str,
                            'timezone': timezone_str,
                            'deadline_type': deadline_type,
                            'description': deadline_context[:500],
                            'is_primary': False,
                        })
                    except (ParserError, ValueError) as e:
                        logger.debug(f"Could not parse date from context: {deadline_context[:100]}")
                        continue
                    break
        
        for deadline in deadlines:
            if deadline['deadline_type'] in ['offers_due', 'submission']:
                deadline['is_primary'] = True
                break
        
        logger.info(f"Extracted {len(deadlines)} deadlines from document")
        return deadlines